#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pandas>=2.0.0",
#   "openpyxl>=3.1.0",
# ]
# ///
"""
Parse Excel files using 4 extraction methods: basic CSV export, structure with JSON,
detailed analysis with statistics, and formula extraction.

Usage: ./parse_xlsx.py <file_path> <output_dir>
"""

import sys
import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook


def parse_method1_pandas(source_file, output_dir):
    """Export all sheets as CSV with metadata using pandas."""
    print("    Method 1: pandas (basic)...")
    method_dir = output_dir / "pandas_basic"
    method_dir.mkdir(exist_ok=True)
    csv_dir = method_dir / "csv"
    csv_dir.mkdir(exist_ok=True)

    try:
        excel_data = pd.read_excel(source_file, sheet_name=None, engine="openpyxl")

        for sheet_name, df in excel_data.items():
            sheet_name_clean = (
                sheet_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
            )
            csv_path = csv_dir / f"{sheet_name_clean}.csv"
            df.to_csv(csv_path, index=False)

        metadata = {
            "sheets": list(excel_data.keys()),
            "source": str(source_file),
            "rows": {name: len(df) for name, df in excel_data.items()},
            "columns": {name: list(df.columns) for name, df in excel_data.items()},
        }

        with open(method_dir / "metadata.json", "w") as f:
            json.dump(metadata, f, indent=2)

        print(f"      ✓ Extracted {len(excel_data)} sheets as CSV")
        return True
    except Exception as e:
        print(f"      ⚠ Failed: {str(e)[:100]}")
        return False


def parse_method2_openpyxl(source_file, output_dir):
    """Extract sheets as JSON with structure metadata using openpyxl."""
    print("    Method 2: openpyxl (structure)...")
    method_dir = output_dir / "openpyxl_structure"
    method_dir.mkdir(exist_ok=True)

    try:
        wb = load_workbook(source_file, data_only=True)

        sheets_info = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets_info[sheet_name] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions": ws.dimensions,
            }

        metadata = {
            "workbook": str(source_file),
            "sheets": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
            "sheet_details": sheets_info,
        }

        with open(method_dir / "workbook_metadata.json", "w") as f:
            json.dump(metadata, f, indent=2)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_name_clean = (
                sheet_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
            )

            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else "" for cell in row])

            with open(method_dir / f"{sheet_name_clean}.json", "w") as f:
                json.dump(data, f, indent=2)

        print(f"      ✓ Extracted {len(wb.sheetnames)} sheets with metadata")
        return True
    except Exception as e:
        print(f"      ⚠ Failed: {str(e)[:100]}")
        return False


def parse_method3_pandas_detailed(source_file, output_dir):
    """Analyze data types, calculate statistics, and count missing values."""
    print("    Method 3: pandas (detailed)...")
    method_dir = output_dir / "pandas_detailed"
    method_dir.mkdir(exist_ok=True)
    stats_dir = method_dir / "statistics"
    stats_dir.mkdir(exist_ok=True)

    try:
        excel_data = pd.read_excel(source_file, sheet_name=None, engine="openpyxl")

        sheets_analysis = {}

        for sheet_name, df in excel_data.items():
            sheet_name_clean = (
                sheet_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
            )

            dtypes_info = {col: str(dtype) for col, dtype in df.dtypes.items()}

            stats = {}
            try:
                numeric_stats = df.describe().to_dict()
                stats["numeric"] = numeric_stats
            except:
                pass

            missing = df.isnull().sum().to_dict()

            sheets_analysis[sheet_name] = {
                "shape": df.shape,
                "columns": list(df.columns),
                "dtypes": dtypes_info,
                "missing_values": {k: int(v) for k, v in missing.items()},
                "statistics": stats,
            }

            with open(stats_dir / f"{sheet_name_clean}_analysis.json", "w") as f:
                json.dump(sheets_analysis[sheet_name], f, indent=2)

        with open(method_dir / "sheets_analysis.json", "w") as f:
            json.dump(sheets_analysis, f, indent=2)

        print(f"      ✓ Analyzed {len(excel_data)} sheets with statistics")
        return True
    except Exception as e:
        print(f"      ⚠ Failed: {str(e)[:100]}")
        return False


def parse_method4_openpyxl_formulas(source_file, output_dir):
    """Extract formulas with cell coordinates (not calculated values)."""
    print("    Method 4: openpyxl (formulas)...")
    method_dir = output_dir / "openpyxl_formulas"
    method_dir.mkdir(exist_ok=True)

    try:
        wb = load_workbook(source_file, data_only=False)

        formulas_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_formulas = []

            for row_idx, row in enumerate(ws.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        sheet_formulas.append(
                            {
                                "cell": cell.coordinate,
                                "row": row_idx,
                                "col": col_idx,
                                "formula": cell.value,
                            }
                        )

            if sheet_formulas:
                formulas_data[sheet_name] = sheet_formulas

        with open(method_dir / "formulas.json", "w") as f:
            json.dump(formulas_data, f, indent=2)

        total_formulas = sum(len(formulas) for formulas in formulas_data.values())
        print(
            f"      ✓ Extracted {total_formulas} formulas from {len(formulas_data)} sheets"
        )
        return True
    except Exception as e:
        print(f"      ⚠ Failed: {str(e)[:100]}")
        return False


def main():
    if len(sys.argv) != 3:
        print("Usage: ./parse_xlsx.py <file_path> <output_dir>")
        sys.exit(1)

    source_file = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])

    if not source_file.exists():
        print(f"Error: File not found: {source_file}")
        sys.exit(1)

    print(f"Parsing XLSX: {source_file.name}")
    print(f"Output directory: {output_dir}")
    print()

    output_dir.mkdir(parents=True, exist_ok=True)

    methods_success = {
        "method1_pandas_basic": parse_method1_pandas(source_file, output_dir),
        "method2_openpyxl_structure": parse_method2_openpyxl(source_file, output_dir),
        "method3_pandas_detailed": parse_method3_pandas_detailed(
            source_file, output_dir
        ),
        "method4_openpyxl_formulas": parse_method4_openpyxl_formulas(
            source_file, output_dir
        ),
    }

    success_count = sum(methods_success.values())
    print(f"\n  ✓ Completed: {success_count}/{len(methods_success)} methods successful")

    results = {
        "timestamp": datetime.now().isoformat(),
        "source_file": str(source_file),
        "output_dir": str(output_dir),
        "methods": methods_success,
        "success_count": success_count,
    }

    with open(output_dir / "parsing_results.json", "w") as f:
        json.dump(results, f, indent=2)

    return 0 if success_count > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
